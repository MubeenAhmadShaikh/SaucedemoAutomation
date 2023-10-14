import openpyxl


class UserCreds:

    user_credentials = [
        {'username': 'standard_user', 'password': 'secret_sauce'},
        {'username': 'standard_user', 'password': 'secr_sauce'},
        {'username': 'stdard_user', 'password': 'secret_sauce'},
        {'username': 'standard_er', 'password': 'secret_sce'},
        {'username': 'locked_out_user', 'password': 'secret_sauce'},
        {'username': 'problem_user', 'password': 'secret_sauce'},
        {'username': 'performance_glitch_user', 'password': 'secret_sauce'},
    ]

    @staticmethod
    def getUserCreds():
        work_book = openpyxl.load_workbook("usercredentials.xlsx")
        sheet = work_book["all"]
        dict = {}
        # for i in range(2, sheet.max_row + 1):
        #     # if sheet.cell(row=i, column=1).value and sheet.cell(row=i, column=1).value.find(user) != -1:
        #     if sheet.cell(row=i, column=1).value == username:
        #         for j in range(2, sheet.max_column + 1):
        #             dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        # return [dict]
        user_data = []
        for row in sheet.iter_rows(min_row=2, max_col=3):
            creds_dict = {
                "username": row[1].value,
                "password": row[2].value
            }
            user_data.append(creds_dict)

        return user_data

    @staticmethod
    def get_invalid_user_credentials():
        work_book = openpyxl.load_workbook(
            "usercredentials.xlsx")
        sheet = work_book["invalid"]
        user_data = []
        for row in sheet.iter_rows(min_row=2, max_col=2):
            creds_dict = {
                "username": row[0].value,
                "password": row[1].value
            }
            user_data.append(creds_dict)

        return user_data

